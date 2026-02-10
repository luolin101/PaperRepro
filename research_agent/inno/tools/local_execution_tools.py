"""
Tools for executing scripts directly on the local machine.
These tools are designed for local environment (LocalEnv) usage.
"""
import subprocess
import os
import tempfile
import re
import shutil
from research_agent.inno.registry import register_tool
from typing import Optional, List


@register_tool("execute_python_code")
def execute_python_code(code: str, cwd: str = None) -> str:
    """
    Execute Python code directly from a string.
    
    Args:
        code: Python code to execute as a string.
        cwd: Working directory for code execution (optional).
    
    Returns:
        A string containing the stdout and stderr output from the code execution.
    """
    try:
        with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False, encoding='utf-8') as f:
            f.write(code)
            temp_path = f.name
        
        try:
            if cwd:
                if not os.path.isabs(cwd):
                    cwd = os.path.abspath(cwd)
                if not os.path.exists(cwd):
                    return f"Error: Working directory does not exist: {cwd}"
            
            cmd = ["python", temp_path]
            result = subprocess.run(
                cmd,
                cwd=cwd,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace"
            )
            
            output = []
            if result.stdout:
                output.append(f"STDOUT:\n{result.stdout}")
            if result.stderr:
                output.append(f"STDERR:\n{result.stderr}")
            output.append(f"Exit code: {result.returncode}")
            
            return "\n".join(output) if output else "Code executed (no output)"
        finally:
            try:
                os.unlink(temp_path)
            except:
                pass
    except Exception as e:
        return f"Error executing Python code: {e}"


@register_tool("run_local_python")
def run_local_python(script_path: str, cwd: str = None, args: list = None) -> str:
    """
    Execute a Python script directly on the local machine.
    
    This tool calls the local Python interpreter to run a script file.
    It does not require Docker or any environment setup.
    
    Args:
        script_path: Absolute or relative path to the Python script to execute.
        cwd: Working directory for the script execution (optional).
        args: Additional command-line arguments to pass to the script (optional).
    
    Returns:
        A string containing the stdout and stderr output from the script execution.
    """
    try:
        # Convert to absolute path if relative
        if not os.path.isabs(script_path):
            script_path = os.path.abspath(script_path)
        
        # Check if script exists
        if not os.path.exists(script_path):
            return f"Error: Script file not found: {script_path}"
        
        # Build command
        cmd = ["python", script_path]
        if args:
            cmd.extend(args)
        
        # Set working directory
        if cwd:
            if not os.path.isabs(cwd):
                cwd = os.path.abspath(cwd)
            if not os.path.exists(cwd):
                return f"Error: Working directory does not exist: {cwd}"
        else:
            cwd = os.path.dirname(script_path)
        
        # Execute the script
        result = subprocess.run(
            cmd,
            cwd=cwd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace"
        )
        
        output = []
        if result.stdout:
            output.append(f"STDOUT:\n{result.stdout}")
        if result.stderr:
            output.append(f"STDERR:\n{result.stderr}")
        output.append(f"Exit code: {result.returncode}")
        
        return "\n".join(output) if output else "Script executed (no output)"
    except Exception as e:
        return f"Error executing Python script: {e}"


@register_tool("run_local_stata")
def run_local_stata(do_file: str, log_file: str = None) -> str:
    """
    Execute a Stata .do file on the local machine using pystata.

    This tool uses `pystata` to call a local Stata installation, wraps the
    target .do file with a small helper .do script that opens a log, runs
    the target, then closes the log. It then returns a short, human-readable
    summary, including a truncated view of the log file.

    Args:
        do_file: Path to the Stata .do file to execute (absolute or relative).
        log_file: Optional path for the Stata log file. If not provided,
                  defaults to `<basename_of_do_file>.log` in the same directory
                  as `do_file`.

    Returns:
        A string summarizing execution status and a truncated preview of the log.
        The full log is saved on disk and its path is included in the message.
    """
    try:
        try:
            import stata_setup
            # Read Stata path from environment variable, default to C:\Soft if not set
            stata_path = os.getenv("STATA_PATH", r"C:\Soft")
            stata_edition = os.getenv("STATA_EDITION", "mp")
            stata_setup.config(stata_path, stata_edition) 
            from pystata import stata  # type: ignore
        except ImportError:
            return (
                "Error: pystata is not installed or not configured.\n"
                "Please install and configure pystata with Stata, e.g.:\n"
                "  pip install pystata\n"
                "and ensure Stata is properly licensed and accessible."
            )

        # Normalize paths
        do_file = os.path.abspath(do_file)
        if not os.path.exists(do_file):
            return f"Error: Stata .do file not found: {do_file}"

        if log_file is None:
            base, _ = os.path.splitext(do_file)
            log_file = base + ".log"
        log_file = os.path.abspath(log_file)

        # Ensure log directory exists
        log_dir = os.path.dirname(log_file)
        if log_dir and not os.path.exists(log_dir):
            os.makedirs(log_dir, exist_ok=True)

        # Create a temporary wrapper .do file that opens a log, runs the target, and closes the log
        with tempfile.NamedTemporaryFile(
            suffix=".do",
            delete=False,
            mode="w",
            encoding="utf-8"
        ) as f:
            f.write("capture log close _all\n")
            f.write(f'log using "{log_file}", replace text\n')
            f.write(f'do "{do_file}"\n')
            f.write("capture log close\n")
            wrapper_do = f.name

        run_error = None
        try:
            # Run the wrapper .do file via pystata
            stata.run(f'do "{wrapper_do}"', echo=False, inline=False)
        except Exception as e:
            run_error = str(e)
        finally:
            try:
                os.unlink(wrapper_do)
            except OSError:
                pass

        # Read and truncate log content if available
        log_preview = ""
        if os.path.exists(log_file):
            try:
                with open(log_file, "r", encoding="utf-8", errors="replace") as lf:
                    content = lf.read()
                max_chars = 1000
                head_chars = 500
                tail_chars = 500
                if len(content) <= max_chars:
                    log_preview = content
                else:
                    head = content[:head_chars]
                    tail = content[-tail_chars:]
                    log_preview = (
                        head
                        + "\n\n... [log truncated] ...\n\n"
                        + tail
                    )
            except Exception as e:
                log_preview = f"(Failed to read log file: {e})"
        else:
            log_preview = "(Log file not found after execution.)"

        # Build return message
        lines = []
        if run_error is None:
            lines.append("✅ Stata .do file executed via pystata without raising an exception.")
        else:
            lines.append("⚠️  Stata execution reported an error via pystata:")
            lines.append(run_error)

        lines.append(f"\nFull execution log saved at:\n  {log_file}")
        lines.append("\nBelow is a truncated preview of the log (head and tail):")
        lines.append("=" * 60)
        lines.append(log_preview)
        lines.append("=" * 60)

        return "\n".join(lines)
    except Exception as e:
        return f"Error executing Stata .do file via pystata: {e}"


@register_tool("run_local_r")
def run_local_r(script_file: str, log_file: str = None) -> str:
    """
    Execute an R script file or R Markdown file on the local machine.

    For .R files: Uses Rscript directly, which captures all output (stdout/stderr) 
    including print(), cat(), warnings(), and errors.
    
    For .Rmd/.rmd files: Uses rmarkdown::render() to render the R Markdown file,
    which will execute all R code chunks and generate the output document.

    The output is saved to a log file, and a truncated preview is returned.

    Args:
        script_file: Path to the R script file (.R) or R Markdown file (.Rmd, .rmd) 
                     to execute (absolute or relative).
        log_file: Optional path for the log file. If not provided,
                  defaults to `<basename_of_script_file>.log` in the same directory
                  as `script_file`.

    Returns:
        A string summarizing execution status and a truncated preview of the log.
        The full log is saved on disk and its path is included in the message.
    """
    try:
        # Normalize paths
        script_file = os.path.abspath(script_file)
        if not os.path.exists(script_file):
            return f"Error: R script file not found: {script_file}"

        if log_file is None:
            base, _ = os.path.splitext(script_file)
            log_file = base + ".log"
        log_file = os.path.abspath(log_file)

        # Ensure log directory exists
        log_dir = os.path.dirname(log_file)
        if log_dir and not os.path.exists(log_dir):
            os.makedirs(log_dir, exist_ok=True)

        script_dir = os.path.dirname(script_file)
        script_name = os.path.basename(script_file)
        
        # Check if file is R Markdown
        file_ext = os.path.splitext(script_file)[1].lower()
        is_rmd = file_ext in ['.rmd', '.rmarkdown']
        
        # Build Rscript command based on file type
        if is_rmd:
            # For R Markdown files, use rmarkdown::render()
            # Escape the file path for use in R command
            script_path_escaped = script_file.replace("\\", "/").replace('"', '\\"')
            r_command = f'rmarkdown::render("{script_path_escaped}")'
            cmd = ["Rscript", "-e", r_command]
        else:
            # For regular R scripts, use Rscript directly
            cmd = ["Rscript", script_name]

        # Determine encoding for R output
        # On Windows, R may use system default encoding (GBK/cp936)
        # On Unix-like systems, usually UTF-8
        if os.name == "nt":  # Windows
            output_encoding = 'gbk'
        else:
            output_encoding = 'utf-8'

        # Execute the script and capture output
        run_error = None
        stdout_text = ""
        stderr_text = ""
        returncode = 0

        try:
            result = subprocess.run(
                cmd,
                cwd=script_dir,  # Set working directory to script directory
                capture_output=True,
                text=False,  # Get bytes first, decode manually
                timeout=3600  # 1 hour timeout
            )
            returncode = result.returncode

            # Decode output with proper encoding
            if result.stdout:
                for enc in [output_encoding, "utf-8", "latin1"]:
                    try:
                        stdout_text = result.stdout.decode(enc, errors="replace")
                        break
                    except UnicodeDecodeError:
                        continue

            if result.stderr:
                for enc in [output_encoding, "utf-8", "latin1"]:
                    try:
                        stderr_text = result.stderr.decode(enc, errors="replace")
                        break
                    except UnicodeDecodeError:
                        continue

        except subprocess.TimeoutExpired:
            run_error = f"{'R Markdown' if is_rmd else 'R script'} execution timed out (exceeded 1 hour)"
        except FileNotFoundError:
            run_error = (
                "Rscript executable not found. Please ensure R is installed and "
                "available in your system PATH (you should be able to run 'Rscript' from command line)."
                + (" Also ensure that the 'rmarkdown' package is installed if running .Rmd files." if is_rmd else "")
            )
        except Exception as e:
            run_error = f"Error executing {'R Markdown file' if is_rmd else 'R script'}: {e}"

        # Combine stdout and stderr for log file
        log_content = ""
        if stdout_text:
            log_content += stdout_text
        if stderr_text:
            if log_content:
                log_content += "\n"
            log_content += stderr_text

        # Save log to file (always save R output)
        try:
            with open(log_file, "w", encoding="utf-8", errors="replace") as lf:
                if run_error:
                    lf.write(f"Execution error: {run_error}\n")
                if log_content:
                    lf.write(log_content)
        except Exception as e:
            # If log write fails, continue with preview generation
            pass

        # Read and truncate log content for preview
        log_preview = ""
        if os.path.exists(log_file):
            try:
                with open(log_file, "r", encoding="utf-8", errors="replace") as lf:
                    content = lf.read()
                max_chars = 1000
                head_chars = 500
                tail_chars = 500
                if len(content) <= max_chars:
                    log_preview = content
                else:
                    head = content[:head_chars]
                    tail = content[-tail_chars:]
                    log_preview = (
                        head
                        + "\n\n... [log truncated] ...\n\n"
                        + tail
                    )
            except Exception as e:
                log_preview = f"(Failed to read log file: {e})"
        else:
            log_preview = "(Log file not found after execution.)"

        # Build return message
        lines = []
        file_type_desc = "R Markdown file" if is_rmd else "R script"
        execution_method = "rmarkdown::render()" if is_rmd else "Rscript"
        
        if run_error:
            lines.append(f"⚠️  {file_type_desc} execution error: {run_error}")
        elif returncode == 0:
            lines.append(f"✅ {file_type_desc} executed successfully via {execution_method}.")
        else:
            lines.append(f"⚠️  {file_type_desc} execution completed with exit code {returncode}.")

        lines.append(f"\nFull execution log saved at:\n  {log_file}")
        lines.append("\nBelow is a truncated preview of the log (head and tail):")
        lines.append("=" * 60)
        lines.append(log_preview)
        lines.append("=" * 60)

        return "\n".join(lines)

    except Exception as e:
        return f"Error executing R script: {e}"


@register_tool("run_local_matlab")
def run_local_matlab(script_file: str, log_file: str = None) -> str:
    """
    Execute a MATLAB .m script file on the local machine using MATLAB command line.

    This tool uses subprocess to call MATLAB directly via command line, which captures
    all output (stdout/stderr) including fprintf, disp, warnings, and errors.
    The output is saved to a log file, and a truncated preview is returned.

    Args:
        script_file: Path to the MATLAB .m script file to execute (absolute or relative).
        log_file: Optional path for the log file. If not provided,
                  defaults to `<basename_of_script_file>.log` in the same directory
                  as `script_file`.

    Returns:
        A string summarizing execution status and a truncated preview of the log.
        The full log is saved on disk and its path is included in the message.
    """
    try:
        # Normalize paths
        script_file = os.path.abspath(script_file)
        if not os.path.exists(script_file):
            return f"Error: MATLAB .m file not found: {script_file}"

        if log_file is None:
            base, _ = os.path.splitext(script_file)
            log_file = base + ".log"
        log_file = os.path.abspath(log_file)

        # Ensure log directory exists
        log_dir = os.path.dirname(log_file)
        if log_dir and not os.path.exists(log_dir):
            os.makedirs(log_dir, exist_ok=True)

        script_dir = os.path.dirname(script_file)
        script_name = os.path.basename(script_file)
        # Remove .m extension if present
        script_base = os.path.splitext(script_name)[0]

        # Build MATLAB command
        # Use -batch flag (available in R2019a+) which runs script and exits
        # -batch automatically sets -nodesktop -nosplash -nodisplay -noFigureWindows
        # Note: script_base should NOT include .m extension for -batch
        cmd = [
            "matlab",
            "-batch",
            script_base
        ]

        # Determine encoding for MATLAB output
        # On Windows, MATLAB typically uses system default encoding (GBK/cp936)
        # On Unix-like systems, usually UTF-8
        if os.name == "nt":  # Windows
            output_encoding = 'gbk'  # Chinese Windows default encoding
        else:
            output_encoding = 'utf-8'

        # Execute the script and capture output
        run_error = None
        stdout_text = ""
        stderr_text = ""
        returncode = 0

        try:
            result = subprocess.run(
                cmd,
                cwd=script_dir,  # Set working directory to script directory
                capture_output=True,
                text=False,  # Get bytes first, decode manually
                timeout=3600  # 1 hour timeout
            )
            returncode = result.returncode

            # Decode output with proper encoding
            if result.stdout:
                for enc in [output_encoding, "utf-8", "latin1"]:
                    try:
                        stdout_text = result.stdout.decode(enc, errors="replace")
                        break
                    except UnicodeDecodeError:
                        continue

            if result.stderr:
                for enc in [output_encoding, "utf-8", "latin1"]:
                    try:
                        stderr_text = result.stderr.decode(enc, errors="replace")
                        break
                    except UnicodeDecodeError:
                        continue

        except subprocess.TimeoutExpired:
            run_error = "MATLAB script execution timed out (exceeded 1 hour)"
        except FileNotFoundError:
            run_error = (
                "MATLAB executable not found. Please ensure MATLAB is installed and "
                "available in your system PATH (you should be able to run 'matlab' from command line)."
            )
        except Exception as e:
            run_error = f"Error executing MATLAB script: {e}"

        # Combine stdout and stderr for log file
        log_content = ""
        if stdout_text:
            log_content += stdout_text
        if stderr_text:
            if log_content:
                log_content += "\n"
            log_content += stderr_text

        # Save log to file (always save MATLAB output)
        try:
            with open(log_file, "w", encoding="utf-8", errors="replace") as lf:
                if run_error:
                    lf.write(f"Execution error: {run_error}\n")
                if log_content:
                    lf.write(log_content)
        except Exception as e:
            # If log write fails, continue with preview generation
            pass

        # Read and truncate log content for preview
        log_preview = ""
        if os.path.exists(log_file):
            try:
                with open(log_file, "r", encoding="utf-8", errors="replace") as lf:
                    content = lf.read()
                max_chars = 1000
                head_chars = 500
                tail_chars = 500
                if len(content) <= max_chars:
                    log_preview = content
                else:
                    head = content[:head_chars]
                    tail = content[-tail_chars:]
                    log_preview = (
                        head
                        + "\n\n... [log truncated] ...\n\n"
                        + tail
                    )
            except Exception as e:
                log_preview = f"(Failed to read log file: {e})"
        else:
            log_preview = "(Log file not found after execution.)"

        # Build return message
        lines = []
        if run_error:
            lines.append(f"⚠️  MATLAB script execution error: {run_error}")
        elif returncode == 0:
            lines.append("✅ MATLAB script executed successfully via MATLAB.")
        else:
            lines.append(f"⚠️  MATLAB script execution completed with exit code {returncode}.")

        lines.append(f"\nFull execution log saved at:\n  {log_file}")
        lines.append("\nBelow is a truncated preview of the log (head and tail):")
        lines.append("=" * 60)
        lines.append(log_preview)
        lines.append("=" * 60)

        return "\n".join(lines)

    except Exception as e:
        return f"Error executing MATLAB script: {e}"


@register_tool("grep_file")
def grep_file(
    file_path: str,
    query: str,
    context_lines: int = 2,
    case_insensitive: bool = True,
    max_matches: int = 50,
) -> str:
    """
    Search for a string/regex in a local file and return matches with surrounding context.

    Args:
        file_path: Path to the file to search (absolute or relative).
        query: Regex pattern (or plain text) to search for.
        context_lines: Number of context lines to include before and after each match.
        case_insensitive: Whether to search case-insensitively.
        max_matches: Maximum number of matches to return (for safety).

    Returns:
        A human-readable string containing matches with line numbers and context.
    """
    try:
        if not os.path.isabs(file_path):
            file_path = os.path.abspath(file_path)
        if not os.path.exists(file_path):
            return f"Error: File not found: {file_path}"

        flags = re.MULTILINE
        if case_insensitive:
            flags |= re.IGNORECASE

        try:
            pattern = re.compile(query, flags)
        except re.error as e:
            return f"Error: Invalid regex pattern: {e}"

        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.read().splitlines()

        matches = []
        for idx, line in enumerate(lines):
            if pattern.search(line):
                matches.append(idx)
                if len(matches) >= max_matches:
                    break

        if not matches:
            return f"No matches found for pattern: {query}\nFile: {file_path}"

        out = []
        out.append(f"Matches for pattern: {query}")
        out.append(f"File: {file_path}")
        out.append(f"Total matches shown: {len(matches)} (max {max_matches})")
        out.append("=" * 60)

        for m in matches:
            start = max(0, m - context_lines)
            end = min(len(lines) - 1, m + context_lines)
            out.append(f"\n--- Match at line {m + 1} ---")
            for i in range(start, end + 1):
                prefix = ">" if i == m else " "
                out.append(f"{prefix} {i + 1:5d}: {lines[i]}")
        out.append("\n" + "=" * 60)
        return "\n".join(out)
    except Exception as e:
        return f"Error searching file: {e}"


@register_tool("replace_in_file")
def replace_in_file(
    file_path: str,
    old_text: str,
    new_text: str,
    replace_all: bool = False,
    context_lines: int = 3,
) -> str:
    """
    Replace content in a local text file by matching an exact substring.

    Args:
        file_path: Path to the file to modify (absolute or relative).
        old_text: Exact source content to replace (must appear in the file).
        new_text: Replacement content.
        replace_all: If True, replace all occurrences; otherwise replace only the first occurrence.
        context_lines: Number of surrounding lines to show before/after the replaced region (for diff preview).

    Returns:
        A human-readable summary of what changed.
    """
    try:
        if not os.path.isabs(file_path):
            file_path = os.path.abspath(file_path)
        if not os.path.exists(file_path):
            return f"Error: File not found: {file_path}"

        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()

        if old_text not in content:
            return (
                "Error: old_text not found in file. No changes made.\n"
                f"File: {file_path}\n"
                f"old_text length: {len(old_text)}"
            )

        # Helper: build a small before/after preview around the replaced region (line-based)
        def _preview_around(text: str, start_idx: int, end_idx: int) -> str:
            lines = text.splitlines()
            # Convert byte index -> line index by counting '\n' before the index
            start_line_idx = text[:start_idx].count("\n")
            end_line_idx = text[:end_idx].count("\n")
            preview_start = max(0, start_line_idx - max(0, context_lines))
            preview_end = min(len(lines) - 1, end_line_idx + max(0, context_lines))
            out_lines = []
            for i in range(preview_start, preview_end + 1):
                out_lines.append(f"{i + 1:5d}: {lines[i]}")
            return "\n".join(out_lines)

        if replace_all:
            first_pos = content.find(old_text)
            before_preview = _preview_around(content, first_pos, first_pos + len(old_text))
            new_content = content.replace(old_text, new_text)
            replaced_count = content.count(old_text)
            # Compute after preview around the first replaced region
            after_preview = _preview_around(new_content, first_pos, first_pos + len(new_text))
        else:
            first_pos = content.find(old_text)
            before_preview = _preview_around(content, first_pos, first_pos + len(old_text))
            new_content = content.replace(old_text, new_text, 1)
            replaced_count = 1
            after_preview = _preview_around(new_content, first_pos, first_pos + len(new_text))

        if new_content == content:
            return f"No changes made (content unchanged).\nFile: {file_path}"

        with open(file_path, "w", encoding="utf-8") as f:
            f.write(new_content)

        return (
            "✅ Replacement applied.\n"
            f"File: {file_path}\n"
            f"Occurrences replaced: {replaced_count}\n"
            f"replace_all: {replace_all}\n"
            "\n--- BEFORE (context preview) ---\n"
            f"{before_preview}\n"
            "\n--- AFTER (context preview) ---\n"
            f"{after_preview}"
        )
    except Exception as e:
        return f"Error replacing content in file: {e}"


@register_tool("copy_file")
def copy_file(src_path: str, dst_path: str, overwrite: bool = True) -> str:
    """
    Copy a local file from src_path to dst_path.

    Args:
        src_path: Source file path (absolute or relative).
        dst_path: Destination file path (absolute or relative).
        overwrite: Whether to overwrite if destination exists.

    Returns:
        A human-readable summary of the copy operation.
    """
    try:
        if not os.path.isabs(src_path):
            src_path = os.path.abspath(src_path)
        if not os.path.isabs(dst_path):
            dst_path = os.path.abspath(dst_path)

        if not os.path.exists(src_path):
            return f"Error: Source file not found: {src_path}"
        if os.path.isdir(src_path):
            return f"Error: Source path is a directory, expected a file: {src_path}"

        dst_dir = os.path.dirname(dst_path)
        if dst_dir and not os.path.exists(dst_dir):
            os.makedirs(dst_dir, exist_ok=True)

        if os.path.exists(dst_path) and not overwrite:
            return f"Error: Destination already exists and overwrite=False: {dst_path}"

        shutil.copy2(src_path, dst_path)
        return f"✅ File copied.\nFrom: {src_path}\nTo:   {dst_path}"
    except Exception as e:
        return f"Error copying file: {e}"


@register_tool("pdf_to_images")
def pdf_to_images(
    pdf_path: str,
    output_dir: str = None,
    image_format: str = "png",
    dpi: int = 200,
    max_pages: int = 50,
) -> str:
    """
    Convert a PDF into images (one image per page) on the local machine.

    Tries PyMuPDF (fitz) first; falls back to pdf2image if available.

    Args:
        pdf_path: Path to the PDF file (absolute or relative).
        output_dir: Directory to save images into. If None, uses `<pdf_dir>/<pdf_basename>_images`.
        image_format: Output image format, e.g. "png" or "jpg".
        dpi: Render DPI (used by pdf2image; for PyMuPDF we approximate via scaling).
        max_pages: Safety cap on number of pages to render.

    Returns:
        A human-readable summary including the output directory and generated image paths.
    """
    try:
        if not os.path.isabs(pdf_path):
            pdf_path = os.path.abspath(pdf_path)
        if not os.path.exists(pdf_path):
            return f"Error: PDF file not found: {pdf_path}"
        if os.path.isdir(pdf_path):
            return f"Error: pdf_path is a directory, expected a file: {pdf_path}"

        if output_dir is None:
            base = os.path.splitext(os.path.basename(pdf_path))[0]
            output_dir = os.path.join(os.path.dirname(pdf_path), f"{base}_images")
        if not os.path.isabs(output_dir):
            output_dir = os.path.abspath(output_dir)
        os.makedirs(output_dir, exist_ok=True)

        fmt = (image_format or "png").lower().lstrip(".")
        if fmt not in {"png", "jpg", "jpeg", "webp"}:
            return f"Error: Unsupported image_format '{image_format}'. Use one of: png, jpg, jpeg, webp."

        generated = []

        # Backend 1: PyMuPDF (fitz)
        try:
            import fitz  # type: ignore

            doc = fitz.open(pdf_path)
            page_count = doc.page_count
            if page_count == 0:
                return f"Error: PDF has 0 pages: {pdf_path}"
            if page_count > max_pages:
                return f"Error: PDF has {page_count} pages which exceeds max_pages={max_pages}."

            # Rough mapping to scale: default PDF is 72 dpi
            zoom = max(0.1, float(dpi) / 72.0)
            matrix = fitz.Matrix(zoom, zoom)

            for i in range(page_count):
                page = doc.load_page(i)
                pix = page.get_pixmap(matrix=matrix, alpha=False)
                out_path = os.path.join(output_dir, f"page_{i+1:03d}.{fmt}")
                pix.save(out_path)
                generated.append(out_path)
            doc.close()

            return (
                "✅ PDF converted to images using PyMuPDF (fitz).\n"
                f"PDF: {pdf_path}\n"
                f"Output dir: {output_dir}\n"
                "Generated files:\n- " + "\n- ".join(generated)
            )
        except ImportError:
            pass
        except Exception as e:
            # If fitz fails for any reason, try fallback
            fitz_error = str(e)
        else:
            fitz_error = None

        # Backend 2: pdf2image
        try:
            from pdf2image import convert_from_path  # type: ignore

            images = convert_from_path(pdf_path, dpi=dpi)
            if len(images) > max_pages:
                return f"Error: PDF renders to {len(images)} pages which exceeds max_pages={max_pages}."

            for i, img in enumerate(images):
                out_path = os.path.join(output_dir, f"page_{i+1:03d}.{fmt}")
                img.save(out_path)
                generated.append(out_path)

            return (
                "✅ PDF converted to images using pdf2image.\n"
                f"PDF: {pdf_path}\n"
                f"Output dir: {output_dir}\n"
                "Generated files:\n- " + "\n- ".join(generated)
            )
        except ImportError:
            install_hint = (
                "Missing dependency for PDF->image conversion.\n"
                "Install one of:\n"
                "- PyMuPDF: `pip install pymupdf`\n"
                "- pdf2image: `pip install pdf2image` (requires Poppler installed and on PATH)\n"
            )
            if "fitz_error" in locals() and fitz_error:
                install_hint += f"\nPyMuPDF attempt failed with: {fitz_error}\n"
            return "Error: Cannot convert PDF to images. " + install_hint
        except Exception as e:
            details = f"pdf2image failed: {e}"
            if "fitz_error" in locals() and fitz_error:
                details = f"PyMuPDF failed: {fitz_error}\n" + details
            return f"Error: PDF conversion failed.\n{details}"

    except Exception as e:
        return f"Error converting PDF to images: {e}"


@register_tool("pdf_page_to_image")
def pdf_page_to_image(
    pdf_path: str,
    page_number: int,
    output_path: str = None,
    image_format: str = "png",
    dpi: int = 200,
) -> str:
    """
    Convert a SINGLE page of a PDF into an image file on the local machine.

    This is useful when you only need one figure/page for comparison in scoring.
    Tries PyMuPDF (fitz) first; falls back to pdf2image if available.

    Args:
        pdf_path: Path to the PDF file (absolute or relative).
        page_number: 1-based page number to render (1 = first page).
        output_path: Path to save the output image. If None, uses
                     `<pdf_dir>/<pdf_basename>_page_<NNN>.<fmt>`.
        image_format: Output image format, e.g. "png" or "jpg".
        dpi: Render DPI (used by pdf2image; for PyMuPDF we approximate via scaling).

    Returns:
        A human-readable summary including the output path.
    """
    try:
        if not os.path.isabs(pdf_path):
            pdf_path = os.path.abspath(pdf_path)
        if not os.path.exists(pdf_path):
            return f"Error: PDF file not found: {pdf_path}"
        if os.path.isdir(pdf_path):
            return f"Error: pdf_path is a directory, expected a file: {pdf_path}"

        if page_number is None or int(page_number) < 1:
            return f"Error: page_number must be >= 1 (got {page_number})"
        page_number = int(page_number)

        fmt = (image_format or "png").lower().lstrip(".")
        if fmt not in {"png", "jpg", "jpeg", "webp"}:
            return f"Error: Unsupported image_format '{image_format}'. Use one of: png, jpg, jpeg, webp."

        if output_path is None:
            base = os.path.splitext(os.path.basename(pdf_path))[0]
            output_path = os.path.join(
                os.path.dirname(pdf_path),
                f"{base}_page_{page_number:03d}.{fmt}",
            )
        if not os.path.isabs(output_path):
            output_path = os.path.abspath(output_path)
        out_dir = os.path.dirname(output_path)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        # Backend 1: PyMuPDF (fitz)
        fitz_error = None
        try:
            import fitz  # type: ignore

            doc = fitz.open(pdf_path)
            page_count = doc.page_count
            if page_number > page_count:
                doc.close()
                return f"Error: page_number {page_number} exceeds PDF page count {page_count}."

            # Rough mapping to scale: default PDF is 72 dpi
            zoom = max(0.1, float(dpi) / 72.0)
            matrix = fitz.Matrix(zoom, zoom)
            page = doc.load_page(page_number - 1)
            pix = page.get_pixmap(matrix=matrix, alpha=False)
            pix.save(output_path)
            doc.close()

            return (
                "✅ PDF page converted to image using PyMuPDF (fitz).\n"
                f"PDF: {pdf_path}\n"
                f"Page: {page_number}\n"
                f"Output: {output_path}"
            )
        except ImportError:
            pass
        except Exception as e:
            fitz_error = str(e)

        # Backend 2: pdf2image
        try:
            from pdf2image import convert_from_path  # type: ignore

            # pdf2image uses 1-based first_page/last_page
            images = convert_from_path(
                pdf_path,
                dpi=dpi,
                first_page=page_number,
                last_page=page_number,
            )
            if not images:
                msg = "Error: pdf2image returned no images."
                if fitz_error:
                    msg += f"\nPyMuPDF failed: {fitz_error}"
                return msg

            images[0].save(output_path)
            return (
                "✅ PDF page converted to image using pdf2image.\n"
                f"PDF: {pdf_path}\n"
                f"Page: {page_number}\n"
                f"Output: {output_path}"
            )
        except ImportError:
            install_hint = (
                "Missing dependency for PDF->image conversion.\n"
                "Install one of:\n"
                "- PyMuPDF: `pip install pymupdf`\n"
                "- pdf2image: `pip install pdf2image` (requires Poppler installed and on PATH)\n"
            )
            if fitz_error:
                install_hint += f"\nPyMuPDF attempt failed with: {fitz_error}\n"
            return "Error: Cannot convert PDF page to image. " + install_hint
        except Exception as e:
            details = f"pdf2image failed: {e}"
            if fitz_error:
                details = f"PyMuPDF failed: {fitz_error}\n" + details
            return f"Error: PDF page conversion failed.\n{details}"

    except Exception as e:
        return f"Error converting PDF page to image: {e}"


def _find_ghostscript_command():
    """
    Find Ghostscript executable on the system.
    On Windows, looks for gswinc64.exe, gswin64.exe, or gs.exe.
    On other systems, looks for gs.
    
    Returns:
        Path to Ghostscript executable, or None if not found.
    """
    # Try different executable names (Windows-specific first)
    gs_names = []
    if os.name == "nt":  # Windows
        gs_names = ["gswinc64.exe", "gswin64.exe", "gswin32.exe", "gs.exe", "gswinc32.exe"]
    else:
        gs_names = ["gs"]
    
    for name in gs_names:
        gs_path = shutil.which(name)
        if gs_path:
            return gs_path
    return None


@register_tool("convert_image_to_png")
def convert_image_to_png(
    image_path: str,
    output_path: str = None,
    dpi: int = 300,
) -> str:
    """
    Convert an image file (e.g., EPS, SVG, TIFF, BMP, etc.) to PNG format.

    This tool attempts multiple conversion methods:
    1. Pillow (PIL) for common formats
    2. ImageMagick (convert command) if available
    3. Ghostscript (gs command) for EPS/PS files if available

    Args:
        image_path: Path to the source image file (absolute or relative).
        output_path: Path for the output PNG file. If None, uses `<image_dir>/<image_basename>.png`.
        dpi: Resolution for vector formats (EPS, SVG) when converting (default: 300).

    Returns:
        A human-readable summary of the conversion result, including the output file path.
    """
    try:
        if not os.path.isabs(image_path):
            image_path = os.path.abspath(image_path)
        if not os.path.exists(image_path):
            return f"Error: Image file not found: {image_path}"
        if os.path.isdir(image_path):
            return f"Error: image_path is a directory, expected a file: {image_path}"

        # Determine output path
        if output_path is None:
            base = os.path.splitext(os.path.basename(image_path))[0]
            output_path = os.path.join(os.path.dirname(image_path), f"{base}.png")
        if not os.path.isabs(output_path):
            output_path = os.path.abspath(output_path)

        # Ensure output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)

        # Get file extension
        _, ext = os.path.splitext(image_path)
        ext_lower = ext.lower()

        # Initialize error variables
        pil_error = "Not attempted"
        magick_error = "Not attempted"
        gs_error = "Not attempted"

        # Method 1: Try Pillow (PIL) first for common formats
        # Note: For EPS files, PIL requires Ghostscript. If Ghostscript is not available,
        # PIL will fail, so we skip PIL for EPS files if Ghostscript is not found.
        try:
            from PIL import Image  # type: ignore

            # For EPS files, check if Ghostscript is available first
            # PIL cannot handle EPS without Ghostscript backend
            if ext_lower == '.eps':
                gs_available = _find_ghostscript_command() is not None
                if not gs_available:
                    pil_error = "EPS files require Ghostscript (gswinc64.exe/gswin64.exe/gs.exe) for PIL/Pillow processing, but Ghostscript not found in PATH"
                else:
                    try:
                        # Try opening EPS with Pillow (requires Ghostscript backend)
                        img = Image.open(image_path)
                        # Convert to RGB if necessary (EPS might be in different mode)
                        if img.mode in ('RGBA', 'LA', 'P'):
                            # Create white background for transparent images
                            rgb_img = Image.new('RGB', img.size, (255, 255, 255))
                            if img.mode == 'P':
                                img = img.convert('RGBA')
                            rgb_img.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                            img = rgb_img
                        elif img.mode != 'RGB':
                            img = img.convert('RGB')
                        img.save(output_path, 'PNG', dpi=(dpi, dpi))
                        return f"✅ Image converted to PNG using Pillow.\nInput:  {image_path}\nOutput: {output_path}"
                    except Exception as e:
                        # Pillow failed, try other methods
                        pil_error = str(e)
            else:
                # For non-EPS formats, try Pillow directly
                try:
                    img = Image.open(image_path)
                    # Convert to RGB if necessary
                    if img.mode in ('RGBA', 'LA'):
                        rgb_img = Image.new('RGB', img.size, (255, 255, 255))
                        rgb_img.paste(img, mask=img.split()[-1])
                        img = rgb_img
                    elif img.mode != 'RGB':
                        img = img.convert('RGB')
                    img.save(output_path, 'PNG')
                    return f"✅ Image converted to PNG using Pillow.\nInput:  {image_path}\nOutput: {output_path}"
                except Exception as e:
                    pil_error = str(e)
        except ImportError:
            pil_error = "Pillow (PIL) not available"
        except Exception as e:
            pil_error = str(e)

        # Method 2: Try ImageMagick (convert command)
        # Check for ImageMagick 7+ first (magick), then ImageMagick 6 (convert)
        magick_cmd = shutil.which("magick")
        convert_cmd = shutil.which("convert")
        
        if magick_cmd:
            # ImageMagick 7+ uses: magick input.eps -density 300 output.png
            try:
                cmd = [magick_cmd, "-density", str(dpi), image_path, output_path]
                result = subprocess.run(
                    cmd,
                    capture_output=True,
                    text=True,
                    timeout=60
                )
                if result.returncode == 0 and os.path.exists(output_path):
                    return f"✅ Image converted to PNG using ImageMagick.\nInput:  {image_path}\nOutput: {output_path}"
                else:
                    magick_error = result.stderr if result.stderr else result.stdout if result.stdout else "Unknown error"
            except subprocess.TimeoutExpired:
                magick_error = "ImageMagick conversion timed out"
            except Exception as e:
                magick_error = str(e)
        elif convert_cmd:
            # ImageMagick 6 uses: convert -density 300 input.eps output.png
            try:
                cmd = [convert_cmd, "-density", str(dpi), image_path, output_path]
                result = subprocess.run(
                    cmd,
                    capture_output=True,
                    text=True,
                    timeout=60
                )
                if result.returncode == 0 and os.path.exists(output_path):
                    return f"✅ Image converted to PNG using ImageMagick.\nInput:  {image_path}\nOutput: {output_path}"
                else:
                    magick_error = result.stderr if result.stderr else result.stdout if result.stdout else "Unknown error"
            except subprocess.TimeoutExpired:
                magick_error = "ImageMagick conversion timed out"
            except Exception as e:
                magick_error = str(e)
        else:
            magick_error = "ImageMagick (magick/convert) not found in PATH"

        # Method 3: Try Ghostscript (for EPS/PS files)
        if ext_lower in ('.eps', '.ps'):
            gs_cmd = _find_ghostscript_command()
            if gs_cmd:
                try:
                    # Ghostscript command to convert EPS to PNG
                    # Note: On Windows, gswinc64.exe/gswin64.exe work the same way for command-line usage
                    cmd = [
                        gs_cmd,
                        "-dNOPAUSE",
                        "-dBATCH",
                        "-sDEVICE=png16m",
                        f"-r{dpi}",
                        f"-sOutputFile={output_path}",
                        image_path
                    ]
                    result = subprocess.run(
                        cmd,
                        capture_output=True,
                        text=True,
                        timeout=60
                    )
                    if result.returncode == 0 and os.path.exists(output_path):
                        return f"✅ Image converted to PNG using Ghostscript.\nInput:  {image_path}\nOutput: {output_path}"
                    else:
                        gs_error = result.stderr if result.stderr else "Unknown error"
                except subprocess.TimeoutExpired:
                    gs_error = "Ghostscript conversion timed out"
                except Exception as e:
                    gs_error = str(e)
            else:
                gs_error = "Ghostscript (gswinc64.exe/gswin64.exe/gs.exe) not found in PATH"

        # If all methods failed, return error with details
        error_msg = f"Error: Failed to convert image to PNG: {image_path}\n"
        error_msg += f"Tried methods:\n"
        error_msg += f"  1. Pillow (PIL): {pil_error}\n"
        error_msg += f"  2. ImageMagick: {magick_error}\n"
        if ext_lower in ('.eps', '.ps'):
            error_msg += f"  3. Ghostscript: {gs_error}\n"
            error_msg += f"\nSuggestions:\n"
            error_msg += f"  - Install Pillow: pip install Pillow\n"
            error_msg += f"  - Install ImageMagick: https://imagemagick.org/script/download.php\n"
            error_msg += f"  - Install Ghostscript: https://www.ghostscript.com/download.html\n"
        else:
            error_msg += f"\nSuggestions:\n"
            error_msg += f"  - Install Pillow: pip install Pillow\n"
            error_msg += f"  - Install ImageMagick: https://imagemagick.org/script/download.php\n"
        
        return error_msg

    except Exception as e:
        return f"Error converting image to PNG: {e}"


@register_tool("xlsx_to_images")
def xlsx_to_images(
    xlsx_path: str,
    output_dir: str = None,
    image_format: str = "png",
    dpi: int = 200,
    max_sheets: int = 50,
    sheet_names: list = None,
) -> str:
    """
    Convert an Excel .xlsx workbook into images (one image per sheet).

    Best-effort behavior:
    - On Windows, if Microsoft Excel is installed, uses COM automation to export each sheet to PDF,
      then converts the PDF page to an image (preserves charts/formatting better).
    - Otherwise falls back to rendering cell values with openpyxl + matplotlib (tables only; charts won't render).

    Args:
        xlsx_path: Path to the .xlsx file (absolute or relative).
        output_dir: Directory to save images into. If None, uses `<xlsx_dir>/<xlsx_basename>_images`.
        image_format: Output image format, e.g. "png" or "jpg".
        dpi: Render DPI (used for PDF->image rendering).
        max_sheets: Safety cap on number of sheets to render.
        sheet_names: Optional list of sheet names to export; if None, exports all sheets.

    Returns:
        A human-readable summary including generated image paths and which backend was used.
    """
    try:
        if not os.path.isabs(xlsx_path):
            xlsx_path = os.path.abspath(xlsx_path)
        if not os.path.exists(xlsx_path):
            return f"Error: XLSX file not found: {xlsx_path}"
        if os.path.isdir(xlsx_path):
            return f"Error: xlsx_path is a directory, expected a file: {xlsx_path}"

        fmt = (image_format or "png").lower().lstrip(".")
        if fmt not in {"png", "jpg", "jpeg", "webp"}:
            return f"Error: Unsupported image_format '{image_format}'. Use one of: png, jpg, jpeg, webp."

        if output_dir is None:
            base = os.path.splitext(os.path.basename(xlsx_path))[0]
            output_dir = os.path.join(os.path.dirname(xlsx_path), f"{base}_images")
        if not os.path.isabs(output_dir):
            output_dir = os.path.abspath(output_dir)
        os.makedirs(output_dir, exist_ok=True)

        # Backend 1: Windows Excel COM -> per-sheet PDF -> image (best fidelity)
        is_windows = os.name == "nt"
        if is_windows:
            try:
                import win32com.client  # type: ignore

                excel = win32com.client.DispatchEx("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                wb = excel.Workbooks.Open(xlsx_path)

                all_sheet_names = [wb.Worksheets(i + 1).Name for i in range(wb.Worksheets.Count)]
                target_names = sheet_names if sheet_names else all_sheet_names
                target_names = [str(n) for n in target_names]

                if len(target_names) > max_sheets:
                    wb.Close(SaveChanges=False)
                    excel.Quit()
                    return f"Error: Requested {len(target_names)} sheets exceeds max_sheets={max_sheets}."

                generated = []
                pdf_paths = []
                for name in target_names:
                    try:
                        ws = wb.Worksheets(name)
                    except Exception:
                        continue

                    safe_name = re.sub(r"[^\w\-_\.]+", "_", name).strip("_") or "sheet"
                    pdf_path = os.path.join(output_dir, f"{safe_name}.pdf")
                    # 0 = PDF format
                    ws.ExportAsFixedFormat(0, pdf_path)
                    pdf_paths.append((name, pdf_path))

                wb.Close(SaveChanges=False)
                excel.Quit()

                # Convert each single-page PDF to an image using existing tool logic (fitz/pdf2image)
                for name, pdf_path in pdf_paths:
                    safe_name = re.sub(r"[^\w\-_\.]+", "_", name).strip("_") or "sheet"
                    out_img = os.path.join(output_dir, f"{safe_name}.{fmt}")
                    # Use page 1 (each exported sheet pdf should be one page)
                    msg = pdf_page_to_image(pdf_path, page_number=1, output_path=out_img, image_format=fmt, dpi=dpi)
                    if msg.startswith("✅"):
                        generated.append(out_img)

                if not generated:
                    return (
                        "Error: Excel COM export ran but produced no images.\n"
                        f"XLSX: {xlsx_path}\nOutput dir: {output_dir}"
                    )

                return (
                    "✅ XLSX converted to images using Windows Excel COM (sheet->PDF->image).\n"
                    f"XLSX: {xlsx_path}\n"
                    f"Output dir: {output_dir}\n"
                    "Generated files:\n- " + "\n- ".join(generated)
                )
            except ImportError:
                pass
            except Exception as e:
                # fall through to openpyxl renderer
                com_error = str(e)
        else:
            com_error = None

        # Backend 2: openpyxl + matplotlib table render (values only)
        try:
            import openpyxl  # type: ignore
            import matplotlib.pyplot as plt  # type: ignore

            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            all_sheet_names = wb.sheetnames
            target_names = sheet_names if sheet_names else all_sheet_names
            target_names = [str(n) for n in target_names if str(n) in all_sheet_names]

            if len(target_names) > max_sheets:
                return f"Error: Requested {len(target_names)} sheets exceeds max_sheets={max_sheets}."

            generated = []
            for name in target_names:
                ws = wb[name]

                # Determine a reasonable used range
                max_row = min(ws.max_row or 1, 200)
                max_col = min(ws.max_column or 1, 50)

                data = []
                for r in range(1, max_row + 1):
                    row = []
                    empty_row = True
                    for c in range(1, max_col + 1):
                        v = ws.cell(row=r, column=c).value
                        s = "" if v is None else str(v)
                        if s != "":
                            empty_row = False
                        row.append(s)
                    if not empty_row:
                        data.append(row)

                safe_name = re.sub(r"[^\w\-_\.]+", "_", name).strip("_") or "sheet"
                out_img = os.path.join(output_dir, f"{safe_name}.{fmt}")

                # Render with matplotlib table
                fig_h = max(3.0, min(30.0, 0.35 * max(1, len(data))))
                fig_w = max(6.0, min(40.0, 0.25 * max(1, max_col)))
                fig, ax = plt.subplots(figsize=(fig_w, fig_h), dpi=dpi)
                ax.axis("off")
                table = ax.table(cellText=data if data else [["(empty sheet)"]], loc="center")
                table.auto_set_font_size(False)
                table.set_fontsize(8)
                table.scale(1, 1.2)
                plt.tight_layout()
                fig.savefig(out_img, bbox_inches="tight")
                plt.close(fig)

                generated.append(out_img)

            if not generated:
                msg = f"Error: No sheets rendered from XLSX: {xlsx_path}"
                if is_windows and com_error:
                    msg += f"\nExcel COM attempt failed: {com_error}"
                return msg

            backend_note = "openpyxl+matplotlib (tables only; charts not preserved)"
            if is_windows and com_error:
                backend_note += f"\nExcel COM attempt failed: {com_error}"

            return (
                f"✅ XLSX converted to images using {backend_note}.\n"
                f"XLSX: {xlsx_path}\n"
                f"Output dir: {output_dir}\n"
                "Generated files:\n- " + "\n- ".join(generated)
            )
        except ImportError:
            hint = (
                "Missing dependency to render XLSX.\n"
                "Install one of:\n"
                "- Best fidelity (Windows + Excel): `pip install pywin32` (requires Microsoft Excel installed)\n"
                "- Fallback (tables only): `pip install openpyxl matplotlib`\n"
            )
            if is_windows and "com_error" in locals() and com_error:
                hint += f"\nExcel COM attempt failed: {com_error}\n"
            return "Error: Cannot convert XLSX to images. " + hint
        except Exception as e:
            msg = f"Error: XLSX conversion failed: {e}"
            if is_windows and "com_error" in locals() and com_error:
                msg += f"\nExcel COM attempt failed: {com_error}"
            return msg

    except Exception as e:
        return f"Error converting XLSX to images: {e}"


if __name__ == "__main__":
    do_file = r"C:\workspace\7\replication_package\APSR Replication Codes.do"
    log_file = r"C:\workspace\7\run_stata.log"
    result = run_local_stata(do_file)
    print(result)
